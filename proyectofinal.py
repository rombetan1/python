from datetime import datetime
from openpyxl import load_workbook
ruta="C:\\Users\\MASTER\\Documents\\0.2 UTP Mintic\\clases\\Seccion 1.4\\ejemplos\\base_crud.xlsx"
def leer(ruta:str, extraer:str):
    archivo_excel = load_workbook(ruta)
    hoja_datos =archivo_excel['Datos del crud']
    hoja_datos = hoja_datos['A2':'F'+str(hoja_datos.max_row)]

    info={}

    for i in hoja_datos:

        if isinstance(i[0].value, int):
            info.setdefault(i[0].value,{'tarea':i[1].value, 'descripcion':i[2].value,
            'estado':i[3].value, 'fecha de inicio':i[4].value,
            'fecha de finalizacion':i[5].value})

    if not(extraer=='todo'):
        info=(filtrar, extraer) 
    
    for i in info:
        print('********* TAREA ********')
        print('Id:'+str(i)+'\n'+'Titulo: '+str(info[i]['tarea'])+'\n'+'Descripcion: '
        +str(info[i]['descripcion']) + '\n'+'Estado: '+str(info[i]['estado'])
        +'\n'+'Fecha creación: '+str(info[i]['fecha de inicio'])
        +'\n'+'fecha de finalizacion: '+str(info[i]['fecha de finalizacion']))
        print()

    return
def filtrar(info:dict, filtro:str):
    aux={}
    for i in info:
        if info[i]['estado']==filtro:
            aux.setdefault(i,info[i])
    return aux

## leer(ruta,'todo')
## inicio de actualizacion en el excel
def actualizar(ruta: str,identificador:int, datos_actualizados:dict):
    archivo_excel = load_workbook(ruta)
    hoja_datos=archivo_excel['Datos del crud']
    hoja_datos = hoja_datos['A2':'F'+str(hoja_datos.max_row)]
    hoja = archivo_excel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizado=6
    encontro=False
    for i in hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True
            for d in datos_actualizados:
                if d=='titulo' and not(datosActualizados[d]==''):
                    hoja.cell(row=fila, column=titulo).value=datosActualizados[d]
                elif d=='descripcion' and not(datosActualizados[d]==''):
                     hoja.cell(row=fila, column=descripcion).value=datosActualizados[d]
                elif d=='estado' and not(datosActualizados[d]==''):
                     hoja.cell(row=fila, column=estado).value=datosActualizados[d]
                elif d=='fecha incio' and not(datosActualizados[d]==''):
                     hoja.cell(row=fila, column=fecha_inicio).value=datosActualizados[d]
                elif d=='fecha finalizacion' and not(datosActualizados[d]==''):
                     hoja.cell(row=fila, column=fecha_finalizado).value=datosActualizados[d]
            archivo_excel.save(ruta)
            if encontro==False:
                print("Error: no existe una tarea con ese identificador")
                print()
            return
            ## inicio de agregar filas al excel
def agregar(ruta:int, datos:dict):
    archivo_excel = load_workbook(ruta)
    hoja_datos=archivo_excel['Datos del crud']
    hoja_datos=hoja_datos['A2':'F'+str(hoja_datos.max_row)]
    hoja=archivo_excel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizacion=6
    for i in hoja_datos:
        if not(isinstance(i[0].value, int)):
            identificador=i[0].row
            hoja.cell(row=identificador, column=1).value=identificador-1
            hoja.cell(row=identificador, column=titulo).value=datos['titulo']
            hoja.cell(row=identificador, column=descripcion).value=datos['descripcion']
            hoja.cell(row=identificador, column=estado).value=datos['estado']
            hoja.cell(row=identificador, column=fecha_inicio).value=datos['fecha inicio']
            hoja.cell(row=identificador, column=fecha_finalizacion).value=datos['fecha finalizacion']
            break
    archivo_excel.save(ruta)
    return
    ## inicio de la funcion borrar acorde al identificador pasado como argumento
def borrar(ruta,identificador):
    archivo_excel=load_workbook(ruta)
    hoja_datos=archivo_excel['Datos del crud']
    hoja_datos=hoja_datos['A2':'F'+str(hoja_datos.max_row)]
    hoja=archivo_excel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizado=6
    encontro=False
    for i in hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True
            hoja.cell(row=fila, column=1).value=""
            hoja.cell(row=fila, column=titulo).value=""
            hoja.cell(row=fila, column=descripcion).value=""
            hoja.cell(row=fila, column=estado).value=""
            hoja.cell(row=fila, column=fecha_inicio).value=""
            hoja.cell(row=fila, column=fecha_finalizado).value=""
    archivo_excel.save(ruta)
    if encontro==False:
        print('Error:No existe una tarea con ese identificador')
        print()
    return
datosActualizados={'titulo':'','descripcion':'','estado':'','fecha inicio':'','fecha finalizacion':''}
while True:
    print("Indique la acción que desea realizar: ")
    print("consultar: 1")
    print("Actualizar: 2")
    print("Crear nueva tarea: 3")
    print("Borrar: 4")
    accion = input("escriba la opcion a seleccionar: ")
    if not(accion=='1') and not(accion=='2')and not(accion=='3')and not(accion=='4'):
        print("valor no valido por favor escriba un numero entre 1 y 4")
    elif accion=='1':
        opc_consulta=''
        print('Indique la tarea que desea consultar:')
        print('todas las tareas:  1')
        print('en espera:  2')
        print('en ejeciucion:   3')
        print('por aprobar:    4')
        print('finalizada:    5')
        opc_consulta = input("escriba la tarea que desea consultar:")
        if opc_consulta=='1':
            print()
            print()
            print("** Consultado todas las tareas***")
            leer(ruta,'todo')
        elif opc_consulta=='2':
            print()
            print()
            print("** Consultado todas las tareas***")
            leer(ruta,'En espera')
        elif opc_consulta=='3':
            print()
            print()
            print("** Consultado todas las tareas***")
            leer(ruta,'En ejecucion')
        elif opc_consulta=='4':
            print()
            print()
            print("** Consultado todas las tareas***")
            leer(ruta,'Por aprobar')
        elif opc_consulta=='5':
            print()
            print()
            print("** Consultado todas las tareas***")
            leer(ruta,'Finalizada')
    elif accion=='2':
        datosActualizados={'titulo':'','descripcion':'','estado':'','fecha inicio':'','fecha finalizado':''}
        print('*****Actualizar tarea*****')
        print()
        id_Actualizar=int(input(' Indique el ID de la tarea que desea actualizar:'))
        print()
        print('****Nuevo titulo*****')
        print('***Nota: si no desea actualizar el titulo solo oprima ENTER')
        datosActualizados['titulo']=input('Indique el nuevo titulo de la tarea:')
        print()
        print('****Nueva descripción*****')
        print('***Nota: si no desea actualizar la descripcion solo oprima ENTER')
        datosActualizados['descripcion']=input('Indique la nueva descripcion de la tarea:')
        print()
        print('****Nuevo estado*****')
        print('En espera   2')
        print('En ejecición   3')
        print('Por aprobar    4')
        print('Finalizada  5')
        print('***Nota: si no desea actualizar el estado solo oprima ENTER')
        estadoNuevo=input('Indique el nuevo estado de la tarea:')
        if estadoNuevo=='2':
            datosActualizados['estado']='En espera'
        elif estadoNuevo=='3':
            datosActualizados['estado']='En ejecucion'
        elif estadoNuevo=='4':
            datosActualizados['estado']='Por aprobar'
        elif estadoNuevo=='5':
            now = datetime.now()
            datosActualizados['estado']='Finalizada'
            datosActualizados['fecha finalizacion']=str(now.day) +'/'+ str(now.month) +'/'+str(now.year)
        
        now = datetime.now()
        datosActualizados['fecha inicio']=str(now.day) +'/'+ str(now.month)+'/'+str(now.year)
        actualizar(ruta,id_Actualizar, datosActualizados)
        print()
    elif accion=='3':
         datosActualizados={'titulo':'','descripcion':'','estado':'','fecha inicio':'','fecha finalizado':''}
         print('***+Crear nueva tarea****')
         print()
         print('****Titulo*****')
         print()
         datosActualizados['titulo']=input('Indique el titulo de la tarea:')
         print()
         print('****Descripcion**')
         datosActualizados['descripcion']=input('Indique la descripcion de la tarea:')
         print()
         datosActualizados['estado']='En espera'
         now=datetime.now()
         datosActualizados['fecha inicio']=str(now.day) +'/'+ str(now.month)+'/'+str(now.year)
         datosActualizados['fecha finalizacion']=''
         agregar(ruta,datosActualizados)
    elif accion=='4':
        print()
        print('***Eliminar Tarea****')
        iden=int(input('Indique el Id de la tarea que desea eliminar: '))
        borrar(ruta,iden)



